#!/usr/bin/env python3
"""
main.py — API web que envuelve resumen_a_excel.py

Recibe un PDF de resumen de tarjeta (BBVA, Banco Nación, Macro) por HTTP
y devuelve el Excel resultante como descarga directa.

Endpoint principal: POST /convertir  (multipart/form-data, campo "file")
"""

import io
import os
import re
import unicodedata

import pdfplumber
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pdfminer.pdfdocument import PDFPasswordIncorrect
from pdfplumber.utils.exceptions import PdfminerException

app = FastAPI(title="Resumen a Excel API")

# --------------------------------------------------------------------------
# Utilidades generales (idéntico al script original)
# --------------------------------------------------------------------------

MESES_ES = {
    "ene": (1, "Enero"), "feb": (2, "Febrero"), "mar": (3, "Marzo"),
    "abr": (4, "Abril"), "may": (5, "Mayo"), "jun": (6, "Junio"),
    "jul": (7, "Julio"), "ago": (8, "Agosto"), "sep": (9, "Septiembre"),
    "set": (9, "Septiembre"), "oct": (10, "Octubre"), "nov": (11, "Noviembre"),
    "dic": (12, "Diciembre"),
}

AMOUNT_RE = re.compile(r"-?\d+(?:\.\d{3})*,\d{2}-?")
CUOTA_RE = re.compile(r"(?:C\.|Cuota\s+)(\d{2}/\d{2})", re.IGNORECASE)
CUPON_RE = re.compile(r"^\d{3,8}[A-Za-z*]{0,3}$")

ADMIN_KEYWORDS = [
    "SALDO ANTERIOR", "SALDO ACTUAL", "SU PAGO EN PESOS", "SU PAGO EN USD",
    "SU PAGO EN DOLARES", "PAGO CAJERO", "TRANSFERENCIA DEUDA", "DEV.IMP",
    "IMPUESTO DE SELLOS", "INTERESES FINANCIACION", "PUNIT. PAG",
    "DB IVA", "DB.IVA", "IIBB PERCEP", "IVA RG", "DB.RG", "PERCEP.",
]

DATE_STYLES = [
    re.compile(r"^\d{2}-[A-Za-zÁ-úñÑ]{3}-\d{2}\b"),
    re.compile(r"^\d{2}\.\d{2}\.\d{2}\b"),
]


def strip_accents(txt):
    return "".join(c for c in unicodedata.normalize("NFD", txt) if unicodedata.category(c) != "Mn")


def parse_amount(token):
    token = token.strip()
    neg = token.startswith("-") or token.endswith("-")
    token = token.strip("-")
    value = float(token.replace(".", "").replace(",", "."))
    return -value if neg else value


def is_admin_line(descripcion):
    up = strip_accents(descripcion.upper())
    return any(strip_accents(k.upper()) in up for k in ADMIN_KEYWORDS)


def titlecase_name(name):
    name = re.sub(r"\s+", " ", name).strip()
    return " ".join(w.capitalize() for w in name.split(" "))


def safe_sheet_name(name, used):
    name = re.sub(r"[:\\/?*\[\]]", "", name).strip()
    name = name[:31] if name else "Hoja"
    base = name
    i = 2
    while name.lower() in used:
        suffix = f" ({i})"
        name = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(name.lower())
    return name


# --------------------------------------------------------------------------
# Extracción de texto del PDF (ahora desde bytes en memoria, no un path)
# --------------------------------------------------------------------------

def extract_lines(pdf_file, password=None):
    lines = []
    with pdfplumber.open(pdf_file, password=password or "") as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            lines.extend(text.split("\n"))
    return lines


# --------------------------------------------------------------------------
# Detección de banco / tarjeta / período
# --------------------------------------------------------------------------

def detect_bank_and_card(full_text):
    up = full_text.upper()
    if "BBVA" in up:
        bank = "BBVA"
    elif "BANCO NACION" in strip_accents(up) or "BNA" in up:
        bank = "BancoNacion"
    elif "MACRO" in up:
        bank = "Macro"
    else:
        bank = "Banco"

    card = ""
    for pattern in [r"VISA\s+GOLD", r"VISA\s+PLATINUM", r"VISA\s+SIGNATURE",
                    r"MASTERCARD\s+GOLD", r"MASTERCARD\s+PLATINUM", r"MASTERCARD\s+BLACK"]:
        m = re.search(pattern, up)
        if m:
            card = re.sub(r"\s+", "", m.group(0).title())
            break
    if not card:
        if "MASTERCARD" in up:
            card = "Mastercard"
        elif "VISA" in up:
            card = "Visa"

    return bank, card


def detect_period(full_text):
    patterns = [
        r"CIERRE ACTUAL[^\d]*?(\d{1,2})[-\s]([A-Za-zÁ-úñÑ]{3})[-\s](\d{2,4})",
        r"Estado de cuenta al\s*:\s*(\d{1,2})[-\s]([A-Za-zÁ-úñÑ]{3})[-\s](\d{2,4})",
    ]
    for pat in patterns:
        m = re.search(pat, full_text, re.IGNORECASE)
        if m:
            dia, mes_abbr, anio = m.groups()
            mes_key = strip_accents(mes_abbr.lower())[:3]
            if mes_key in MESES_ES:
                mes_num, mes_nombre = MESES_ES[mes_key]
                anio = int(anio)
                if anio < 100:
                    anio += 2000
                return anio, mes_num, mes_nombre
    return None, None, None


# --------------------------------------------------------------------------
# Parseo de renglones de transacciones
# --------------------------------------------------------------------------

def match_date_prefix(line):
    for rgx in DATE_STYLES:
        m = rgx.match(line.strip())
        if m:
            return m.group(0), line.strip()[m.end():].strip()
    return None, None


def split_amounts(resto):
    matches = list(AMOUNT_RE.finditer(resto))
    if not matches:
        return None
    cupon_hint = None
    if len(matches) >= 2:
        pre_text = resto[:matches[-2].start()]
        gap = resto[matches[-2].end():matches[-1].start()].strip()
        if re.search(r"(USD|U\$S)\s*$", pre_text, re.IGNORECASE):
            pesos, dolares = 0.0, parse_amount(matches[-1].group(0))
            cut_at = matches[-1].start()
        elif gap and CUPON_RE.match(gap):
            cupon_hint = gap
            pesos, dolares = 0.0, parse_amount(matches[-1].group(0))
            cut_at = matches[-1].start()
        else:
            dolares = parse_amount(matches[-1].group(0))
            pesos = parse_amount(matches[-2].group(0))
            cut_at = matches[-2].start()
    else:
        amt = parse_amount(matches[-1].group(0))
        cut_at = matches[-1].start()
        if re.search(r"\bUS\$|\bU\$S|\bUSD\b", resto, re.IGNORECASE):
            pesos, dolares = 0.0, amt
        else:
            pesos, dolares = amt, 0.0
    texto_antes = resto[:cut_at].strip()
    return texto_antes, pesos, dolares, cupon_hint


def extract_cuota(texto):
    m = CUOTA_RE.search(texto)
    if not m:
        return texto.strip(), ""
    cuota = m.group(1)
    texto = (texto[:m.start()] + texto[m.end():]).strip()
    texto = re.sub(r"\s{2,}", " ", texto)
    return texto.strip(), cuota


def extract_leading_cupon(texto):
    partes = texto.split(" ", 1)
    if len(partes) == 2 and CUPON_RE.match(partes[0]):
        return partes[1].strip(), partes[0]
    if len(partes) == 1 and CUPON_RE.match(partes[0]):
        return "", partes[0]
    return texto, ""


def extract_trailing_cupon(texto):
    partes = texto.rsplit(" ", 1)
    if len(partes) == 2 and CUPON_RE.match(partes[1]):
        return partes[0].strip(), partes[1]
    return texto, ""


def parse_line(line, cupon_style):
    fecha, resto = match_date_prefix(line)
    if fecha is None:
        return None
    result = split_amounts(resto)
    if result is None:
        return None
    texto, pesos, dolares, cupon_hint = result
    if not texto:
        return None

    if cupon_hint:
        cupon = cupon_hint
        if texto.endswith(cupon_hint):
            texto = texto[: -len(cupon_hint)].strip()
        texto, cuota = extract_cuota(texto)
    elif cupon_style == "leading":
        texto, cupon = extract_leading_cupon(texto)
        texto, cuota = extract_cuota(texto)
    else:
        texto, cuota = extract_cuota(texto)
        texto, cupon = extract_trailing_cupon(texto)

    descripcion = re.sub(r"\s{2,}", " ", texto).strip(" -")
    if not descripcion:
        return None
    for rgx in DATE_STYLES:
        if rgx.match(descripcion):
            return None

    return {
        "fecha": fecha,
        "descripcion": descripcion,
        "cuota": cuota,
        "cupon": cupon,
        "pesos": pesos,
        "dolares": dolares,
    }


# --------------------------------------------------------------------------
# Estrategias de agrupado por persona
# --------------------------------------------------------------------------

def group_bbva(lines):
    personas = {}
    otros = []
    seccion_actual = None
    buffer = []

    for raw in lines:
        line = raw.strip()
        if not line:
            continue

        m_inicio = re.match(r"^Consumos\s+(.+)$", line)
        if m_inicio and not line.upper().startswith("TOTAL"):
            if seccion_actual:
                personas.setdefault(seccion_actual, []).extend(buffer)
            seccion_actual = titlecase_name(m_inicio.group(1))
            buffer = []
            continue

        if re.match(r"^TOTAL CONSUMOS DE\s+", line, re.IGNORECASE):
            if seccion_actual:
                personas.setdefault(seccion_actual, []).extend(buffer)
            seccion_actual = None
            buffer = []
            continue

        row = parse_line(line, cupon_style="trailing")
        if row is None:
            continue
        if is_admin_line(row["descripcion"]):
            otros.append(row)
            continue
        if seccion_actual:
            buffer.append(row)
        else:
            otros.append(row)

    return personas, otros


def group_by_trailing_marker(lines, marker_regex):
    personas = {}
    otros = []
    buffer = []

    for raw in lines:
        line = raw.strip()
        if not line:
            continue

        m = marker_regex.match(line)
        if m:
            nombre_bruto = m.group(1)
            amt_matches = list(AMOUNT_RE.finditer(nombre_bruto))
            if amt_matches:
                nombre_bruto = nombre_bruto[: amt_matches[0].start()]
            nombre = titlecase_name(nombre_bruto)
            personales, admins = [], []
            for row in buffer:
                (admins if is_admin_line(row["descripcion"]) else personales).append(row)
            personas.setdefault(nombre, []).extend(personales)
            otros.extend(admins)
            buffer = []
            continue

        row = parse_line(line, cupon_style="leading")
        if row is None:
            continue
        buffer.append(row)

    for row in buffer:
        otros.append(row)

    return personas, otros


def group_nacion_mastercard(lines):
    personas = {}
    otros = []
    buffer = []
    marker = re.compile(r"^TOTAL TITULAR\s+(.+)$", re.IGNORECASE)

    for raw in lines:
        line = raw.strip()
        if not line:
            continue
        m = marker.match(line)
        if m:
            nombre_bruto = m.group(1)
            amt_matches = list(AMOUNT_RE.finditer(nombre_bruto))
            if amt_matches:
                nombre_bruto = nombre_bruto[: amt_matches[0].start()]
            nombre = titlecase_name(nombre_bruto)
            personales, admins = [], []
            for row in buffer:
                (admins if is_admin_line(row["descripcion"]) else personales).append(row)
            personas.setdefault(nombre, []).extend(personales)
            otros.extend(admins)
            buffer = []
            continue

        row = parse_line(line, cupon_style="trailing")
        if row is None:
            continue
        buffer.append(row)

    for row in buffer:
        otros.append(row)

    return personas, otros


# --------------------------------------------------------------------------
# Escritura del Excel
# --------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", size=10, bold=True)

COLUMNS = [
    ("Fecha", 12), ("Descripción", 42), ("Cuota", 10),
    ("Cupón", 12), ("Pesos", 14), ("Dólares", 12),
]


def write_sheet(ws, rows):
    for col, (titulo, ancho) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = ancho

    total_pesos = total_dolares = 0.0
    r = 2
    for row in rows:
        ws.cell(row=r, column=1, value=row["fecha"]).font = BODY_FONT
        ws.cell(row=r, column=2, value=row["descripcion"]).font = BODY_FONT
        ws.cell(row=r, column=3, value=row["cuota"]).font = BODY_FONT
        ws.cell(row=r, column=4, value=row["cupon"]).font = BODY_FONT
        c5 = ws.cell(row=r, column=5, value=row["pesos"])
        c5.font, c5.number_format = BODY_FONT, "#,##0.00"
        c6 = ws.cell(row=r, column=6, value=row["dolares"] or None)
        c6.font, c6.number_format = BODY_FONT, "#,##0.00"
        total_pesos += row["pesos"]
        total_dolares += row["dolares"]
        r += 1

    ws.cell(row=r, column=2, value="TOTAL").font = BOLD_FONT
    c5 = ws.cell(row=r, column=5, value=total_pesos)
    c5.font, c5.number_format = BOLD_FONT, "#,##0.00"
    c6 = ws.cell(row=r, column=6, value=total_dolares or None)
    c6.font, c6.number_format = BOLD_FONT, "#,##0.00"
    ws.freeze_panes = "A2"


def build_workbook(personas, otros):
    wb = Workbook()
    wb.remove(wb.active)
    used = set()

    for nombre, rows in personas.items():
        if not rows:
            continue
        ws = wb.create_sheet(safe_sheet_name(nombre, used))
        write_sheet(ws, rows)

    if otros:
        ws = wb.create_sheet(safe_sheet_name("Otros (pagos e impuestos)", used))
        write_sheet(ws, otros)

    if not wb.sheetnames:
        wb.create_sheet("Sin datos")

    return wb


# --------------------------------------------------------------------------
# Orquestación (ahora trabaja 100% en memoria, sin tocar disco)
# --------------------------------------------------------------------------

def process_pdf_bytes(pdf_bytes, original_filename, password=None):
    pdf_file = io.BytesIO(pdf_bytes)
    lines = extract_lines(pdf_file, password=password)
    full_text = "\n".join(lines)

    bank, card = detect_bank_and_card(full_text)
    anio, mes_num, mes_nombre = detect_period(full_text)

    if bank == "BBVA":
        personas, otros = group_bbva(lines)
    elif bank == "BancoNacion" and "MASTERCARD" in full_text.upper():
        personas, otros = group_nacion_mastercard(lines)
    elif bank in ("BancoNacion", "Macro"):
        marker = re.compile(r"^(?:TARJETA|Tarjeta)\s+\d+\s+Total\s+Consumos\s+de\s+(.+)$")
        personas, otros = group_by_trailing_marker(lines, marker)
    else:
        marker = re.compile(r"^(?:TARJETA|Tarjeta)\s+\d+\s+Total\s+Consumos\s+de\s+(.+)$")
        personas, otros = group_by_trailing_marker(lines, marker)
        if not personas:
            personas, otros = group_bbva(lines)

    nombre_partes = [bank]
    if card:
        nombre_partes.append(card)
    if mes_nombre and anio:
        nombre_partes.append(f"{mes_nombre}{anio}")
    else:
        nombre_partes.append(os.path.splitext(original_filename)[0])

    nombre_archivo = "_".join(nombre_partes) + ".xlsx"
    nombre_archivo = re.sub(r"[^\w\-.]", "_", nombre_archivo)

    wb = build_workbook(personas, otros)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return nombre_archivo, output


# --------------------------------------------------------------------------
# Endpoints
# --------------------------------------------------------------------------

@app.get("/")
def health_check():
    return {"status": "ok"}


@app.post("/convertir")
async def convertir(file: UploadFile = File(...), password: str = Form(None)):
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="El archivo debe ser un PDF")

    pdf_bytes = await file.read()
    if not pdf_bytes:
        raise HTTPException(status_code=400, detail="El archivo está vacío")

    try:
        nombre_archivo, excel_bytes = process_pdf_bytes(pdf_bytes, file.filename, password=password)
    except PdfminerException as e:
        # Si la causa es contraseña faltante/incorrecta, devolvemos un error
        # específico (401) para que la app pueda pedirle la clave al usuario.
        if e.args and isinstance(e.args[0], PDFPasswordIncorrect):
            if password:
                raise HTTPException(status_code=401, detail="password_incorrect")
            raise HTTPException(status_code=401, detail="password_required")
        raise HTTPException(status_code=500, detail=f"Error al leer el PDF: {e}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al procesar el PDF: {e}")

    return StreamingResponse(
        excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre_archivo}"'},
    )
